"""
Looker Studio에서 읽어온 주간 AS 퍼널 데이터를 Excel에 추가하고 대시보드를 재생성합니다.

사용법:
  python update_weekly_funnel.py --period "5.11~5.17" --reg 2969 --view 1334 --click 401 --select 299 --symptom 190 --info 175 --done 114

매주 월요일, Claude를 통해 Looker Studio에서 전주 데이터를 읽어온 뒤 실행합니다.
"""

import argparse
import subprocess
import sys
import json
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(r"H:\내 드라이브\4. MOT 문제해결\2. AS 데이터 관리")
NOTIFICATIONS_PATH = BASE_DIR / "notifications.json"
EXCEL_PATH = BASE_DIR / "AS 현황_시디즈.xlsx"
SHEET_NAME = "AS 페이지, 컨택센터"
GENERATE_SCRIPT = BASE_DIR / "generate_dashboard.py"

# Excel 행 번호 (1-indexed)
ROW_PERIOD   = 15   # 기간
ROW_REG      = 16   # 제품 등록 수
ROW_VIEW     = 17   # 1. AS 신청 메인 페이지 조회
ROW_CLICK    = 18   # 2. AS 신청 버튼 클릭
ROW_SELECT   = 19   # 3. 제품 선택
ROW_SYMPTOM  = 20   # 4. 제품 증상 입력
ROW_INFO     = 21   # 5. 고객 정보 입력
ROW_DONE     = 22   # 6. AS 신청완료
ROW_RATE12   = 23   # 1→2 유입률
ROW_RATE34   = 24   # 3→4 유입률
ROW_RATE45   = 25   # 4→5 유입률


def find_next_col(ws):
    """기간 행에서 첫 번째 빈 열(1-indexed) 반환"""
    for col in range(2, 30):
        if ws.cell(row=ROW_PERIOD, column=col).value is None:
            return col
    raise ValueError("빈 열을 찾을 수 없습니다")


def add_weekly_data(period: str, reg: int, view: int, click: int, select: int, symptom: int, info: int, done: int):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    col = find_next_col(ws)

    rate12 = round(click / view, 9) if view else 0
    rate34 = round(symptom / select, 9) if select else 0
    rate45 = round(info / symptom, 9) if symptom else 0

    ws.cell(row=ROW_PERIOD,  column=col).value = period
    ws.cell(row=ROW_REG,     column=col).value = reg
    ws.cell(row=ROW_VIEW,    column=col).value = view
    ws.cell(row=ROW_CLICK,   column=col).value = click
    ws.cell(row=ROW_SELECT,  column=col).value = select
    ws.cell(row=ROW_SYMPTOM, column=col).value = symptom
    ws.cell(row=ROW_INFO,    column=col).value = info
    ws.cell(row=ROW_DONE,    column=col).value = done
    ws.cell(row=ROW_RATE12,  column=col).value = rate12
    ws.cell(row=ROW_RATE34,  column=col).value = rate34
    ws.cell(row=ROW_RATE45,  column=col).value = rate45

    wb.save(EXCEL_PATH)
    print(f"✓ Excel 업데이트 완료: {period} → {col}열")
    print(f"  제품 등록: {reg:,}  조회: {view:,}  클릭: {click:,}  선택: {select:,}  증상: {symptom:,}  정보: {info:,}  완료: {done:,}")
    print(f"  1→2 유입률: {rate12:.1%}  3→4 유입률: {rate34:.1%}  4→5 유입률: {rate45:.1%}")


def add_notification(period: str, view: int, done: int, conv12: float, conv45: float):
    notifs = json.loads(NOTIFICATIONS_PATH.read_text(encoding="utf-8")) if NOTIFICATIONS_PATH.exists() else []
    notifs.insert(0, {
        "text": f"<b>주간 퍼널 데이터 ({period})</b> 업데이트 완료 — "
                f"메인 조회 {view:,}회 · AS 신청완료 {done:,}건 · "
                f"1→2 유입률 {conv12:.1f}% · 4→5 유입률 {conv45:.1f}%",
        "time": date.today().strftime("%Y.%m.%d"),
        "read": False
    })
    NOTIFICATIONS_PATH.write_text(json.dumps(notifs, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ 알림 기록 완료")


def regenerate_dashboard():
    print("대시보드 재생성 중...")
    result = subprocess.run(
        [sys.executable, str(GENERATE_SCRIPT)],
        capture_output=True, text=True, encoding="utf-8"
    )
    if result.returncode == 0:
        print("✓ AS_대시보드.html 재생성 완료")
    else:
        print(f"⚠ 대시보드 재생성 실패: {result.stderr[:200]}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="주간 AS 퍼널 데이터 업데이트")
    parser.add_argument("--period",  required=True, help='기간 (예: "5.11~5.17")')
    parser.add_argument("--reg",     required=True, type=int, help="정품등록 완료 수")
    parser.add_argument("--view",    required=True, type=int, help="AS 메인 페이지 조회 수")
    parser.add_argument("--click",   required=True, type=int, help="AS 신청 버튼 클릭 수")
    parser.add_argument("--select",  required=True, type=int, help="제품 선택 수")
    parser.add_argument("--symptom", required=True, type=int, help="제품 증상 입력 수")
    parser.add_argument("--info",    required=True, type=int, help="고객 정보 입력 수")
    parser.add_argument("--done",    required=True, type=int, help="AS 신청완료 수")
    parser.add_argument("--no-dashboard", action="store_true", help="대시보드 재생성 건너뜀")
    args = parser.parse_args()

    add_weekly_data(args.period, args.reg, args.view, args.click, args.select, args.symptom, args.info, args.done)

    rate12 = round(args.click / args.view * 100, 1) if args.view else 0
    rate45 = round(args.info / args.symptom * 100, 1) if args.symptom else 0
    add_notification(args.period, args.view, args.done, rate12, rate45)

    if not args.no_dashboard:
        regenerate_dashboard()
